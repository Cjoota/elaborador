import os
import sys
from datetime import datetime
from openpyxl import load_workbook

class ExcelProtocol:
    def __init__(self, nome_arquivo_modelo="modelo.xlsx"):
        # --- CONFIGURAÇÃO ---
        self.pasta_modelos = "modelo_relacoes"
        self.modelo_excel_path = os.path.join(self.pasta_modelos, nome_arquivo_modelo)
        
        self.exportados_path = "exportados_ms"
        os.makedirs(self.exportados_path, exist_ok=True)

    def get_app_path(filename):
        """Retorna o caminho onde o .exe está rodando (e não a pasta temp)"""
        if getattr(sys, 'frozen', False):
            # Se estiver rodando como .exe
            application_path = os.path.dirname(sys.executable)
        else:
            # Se estiver rodando como script normal
            application_path = os.path.dirname(os.path.abspath(__file__))
        
        return os.path.join(application_path, filename)
    
    def registrar_colaborador(self, colaborator: str, cpf: str):
        """
        Lógica de Paginação (Limite D27):
        1. Procura 'Protocolo_MSBOI_Pagina_X.xlsx'.
        2. Varre da linha 8 até a 27.
        3. Se achar vaga, escreve e atualiza a data.
        4. Se chegar na 27 e estiver cheia, fecha e tenta a próxima Página.
        """
        try:
            print(f"[EXCEL] Iniciando registro para: {colaborator}")
            
            agora = datetime.now()
            data_formatada = agora.strftime("%d/%m/%Y")
            
            pagina_atual = 1
            
            while True:
                nome_arquivo = f"Protocolo_MSBOI_Pagina_{pagina_atual}.xlsx"
                caminho_final = os.path.join(self.exportados_path, nome_arquivo)
                
                # --- CENÁRIO 1: PÁGINA NOVA (Não existe ainda) ---
                if not os.path.exists(caminho_final):
                    print(f"[EXCEL] Página {pagina_atual} não existe. Criando nova...")
                    
                    if not os.path.exists(self.modelo_excel_path):
                        print(f"[ERRO CRÍTICO] Modelo não encontrado!")
                        return False

                    wb = load_workbook(self.modelo_excel_path)
                    ws = wb.active
                    
                    # Como é nova, escreve direto na primeira linha (8)
                    ws['H5'] = data_formatada
                    ws['D8'] = colaborator
                    ws['K8'] = cpf
                    
                    wb.save(caminho_final)
                    print(f"[SUCESSO] Criado {nome_arquivo} e salvo na linha 8.")
                    return True

                # --- CENÁRIO 2: PÁGINA JÁ EXISTE (Vamos ver se cabe) ---
                wb = load_workbook(caminho_final)
                ws = wb.active
                
                linha_para_escrever = None

                # Verifica EXATAMENTE da linha 8 até a 27
                # range(8, 28) significa: 8, 9, 10 ... até 27. (O 28 fica de fora)
                for linha in range(8, 28):
                    celula = ws[f'D{linha}']
                    if not celula.value: # Se estiver vazia
                        linha_para_escrever = linha
                        break # Achamos! Para de procurar.
                
                if linha_para_escrever:
                    # --- TEM VAGA NA PÁGINA ATUAL ---
                    ws['H5'] = data_formatada # Atualiza a data do cabeçalho
                    
                    ws[f'D{linha_para_escrever}'] = colaborator
                    ws[f'K{linha_para_escrever}'] = cpf
                    
                    wb.save(caminho_final)
                    print(f"[SUCESSO] Salvo em {nome_arquivo} na linha {linha_para_escrever}.")
                    return True
                
                else:
                    # --- NÃO TEM VAGA (Da 8 até a 27 estava tudo cheio) ---
                    print(f"[AVISO] {nome_arquivo} está completa (até linha 27). Indo para próxima página...")
                    pagina_atual += 1
                    wb.close()
                    # O Loop recomeça, agora procurando a Pagina_2...

        except Exception as e:
            print(f"[ERRO EXCEL] {e}")
            return False